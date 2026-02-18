import os
import io
import json
import asyncio
import logging
import httpx
import openpyxl
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

app = FastAPI(title="LinkedIn Email Enrichment Service")

CARGO_API_URL = "https://api.getcargo.io/v1/tools/017fd330-fc34-42a5-b608-25897c94ba28/execute"
CARGO_API_KEY = os.environ.get("CARGO_API_KEY", "032c2189a0ae75dca4c022d4c99d7a96348f4fcbb8fcfd27d24ee86307382b32")
REQUEST_DELAY_S = float(os.environ.get("REQUEST_DELAY_S", "1.5"))


async def call_cargo(linkedin_url: str) -> str | None:
    """Call getcargo.io asynchronously and return email string, or None on failure."""
    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            resp = await client.post(
                CARGO_API_URL,
                headers={
                    "Authorization": f"Bearer {CARGO_API_KEY}",
                    "Content-Type": "application/json",
                },
                json={"linkedinUrl": linkedin_url},
            )
            resp.raise_for_status()
            data = resp.json()

        logger.info(f"Cargo response for {linkedin_url}: {json.dumps(data)}")

        # Handle various response shapes
        if isinstance(data.get("email"), str) and "@" in data["email"]:
            return data["email"]
        if isinstance(data.get("data"), dict) and "@" in str(data["data"].get("email", "")):
            return data["data"]["email"]
        if isinstance(data.get("result"), dict) and "@" in str(data["result"].get("email", "")):
            return data["result"]["email"]
        if isinstance(data.get("output"), str) and "@" in data["output"]:
            return data["output"]
        if isinstance(data.get("output"), dict) and "@" in str(data["output"].get("email", "")):
            return data["output"]["email"]
        return None

    except Exception as e:
        logger.error(f"Cargo API error for {linkedin_url}: {e}")
        return None


@app.get("/")
def health():
    return {"status": "ok"}


@app.post("/enrich")
async def enrich(file: UploadFile = File(...)):
    if not file.filename.endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported.")

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    try:
        url_col = headers.index("LinkedIn URL") + 1
    except ValueError:
        raise HTTPException(status_code=400, detail="Column 'LinkedIn URL' not found in sheet.")

    if "Email" not in headers:
        email_col = len(headers) + 1
        ws.cell(row=1, column=email_col, value="Email")
    else:
        email_col = headers.index("Email") + 1

    total = ws.max_row - 1
    enriched = 0

    for row_idx in range(2, ws.max_row + 1):
        linkedin_url = ws.cell(row=row_idx, column=url_col).value
        existing_email = ws.cell(row=row_idx, column=email_col).value

        if not linkedin_url or existing_email:
            continue

        logger.info(f"[{row_idx - 1}/{total}] Enriching {linkedin_url}")
        email = await call_cargo(str(linkedin_url).strip())
        ws.cell(row=row_idx, column=email_col, value=email or "")
        enriched += 1
        await asyncio.sleep(REQUEST_DELAY_S)

    logger.info(f"Done. Enriched {enriched}/{total} rows.")

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    filename = file.filename.replace(".xlsx", "_enriched.xlsx")
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
